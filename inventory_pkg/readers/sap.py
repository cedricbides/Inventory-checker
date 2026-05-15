"""
SAP inventory reader.

Handles two file formats automatically:
  Standard : SAP export with Article, Site, Unrestricted Stock columns.
  Simple   : User-modified upload with just SKU + Stock columns.
             SKU -> normalised as 'Article', Stock -> normalised as 'Unrestricted'.
             Site/storage filters are skipped for simple-format files.

Public API
----------
read_sap_file(path, site_filter=None, storage_loc_filter=None)  -> (data, headers)
read_sap_files(paths, site_filter=None, storage_loc_filter=None) -> (merged_data, headers)
scan_sap_columns(paths)                                          -> list[str]
"""
import os

from ..utils import is_valid_sku, to_num, merge_row

def scan_sap_columns(sap_paths):
    """Return a deduplicated ordered header list from SAP files."""
    cols, seen = [], set()
    for path in sap_paths:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                raw_hdrs = [str(h).strip() if h else "" for h in row]
                hdrs = [h for h in raw_hdrs if h]
                if not hdrs:
                    continue
                hdrs_up = {h.upper() for h in hdrs}
                if "SKU" in hdrs_up and "STOCK" in hdrs_up and "ARTICLE" not in hdrs_up:
                    normalised = ["Article", "Unrestricted"] + [
                        h for h in hdrs if h.upper() not in ("SKU", "STOCK")
                    ]
                    for h in normalised:
                        if h and h not in seen:
                            seen.add(h)
                            cols.append(h)
                    print(f"    SAP scan: '{os.path.basename(path)}' is simple "
                          f"SKU/Stock format — normalised to Article/Unrestricted")
                else:
                    for h in hdrs:
                        if h and h not in seen:
                            seen.add(h)
                            cols.append(h)
                break
            wb.close()
        except Exception as e:
            print(f"    SAP scan error {path}: {e}")
    return cols

def read_sap_file(path, site_filter=None, storage_loc_filter=None):
    """
    Read all columns from a SAP xlsx file.

    Returns (data, headers) where data is {article: {col: val}}.
    """
    data, all_headers = {}, []

    def _get_ci(rd_ci, *keys):
        for k in keys:
            if k in rd_ci:
                return rd_ci[k]
        return ""

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        headers = None
        is_simple = False
        diag_sites = set()
        diag_slocs = set()
        diag_count = 0

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                raw_hdrs = [h for h in headers if h]
                hdrs_up = {h.upper() for h in raw_hdrs}

                if "SKU" in hdrs_up and "STOCK" in hdrs_up and "ARTICLE" not in hdrs_up:
                    is_simple = True
                    all_headers = ["Article", "Unrestricted"] + [
                        h for h in raw_hdrs if h.upper() not in ("SKU", "STOCK")
                    ]
                    print(f"    SAP '{os.path.basename(path)}': simple SKU/Stock format "
                          f"— site/storage filters skipped")
                else:
                    all_headers = raw_hdrs
                continue

            if not any(row):
                continue

            row_dict = dict(zip(headers, row))

            if is_simple:
                sku_key = next((k for k in headers if k.upper() == "SKU"), "SKU")
                stock_key = next((k for k in headers if k.upper() == "STOCK"), "Stock")
                article = str(row_dict.get(sku_key, "")).strip()
                if not article or article == "None":
                    continue
                stock_val = to_num(row_dict.get(stock_key, 0))
                clean = {"Article": article, "Unrestricted": stock_val}
                for k, v in row_dict.items():
                    if k and k.upper() not in ("SKU", "STOCK"):
                        clean[k] = str(v).strip() if v is not None else ""
                if article in data:
                    data[article]["Unrestricted"] = (
                        to_num(data[article].get("Unrestricted", 0)) + stock_val
                    )
                else:
                    data[article] = clean
                continue

            row_dict_ci = {str(k).strip().lower(): v for k, v in row_dict.items() if k}
            article = str(row_dict.get("Article", "")).strip()
            site = str(row_dict.get("Site", "")).strip()
            storage_loc = str(_get_ci(row_dict_ci,
                "storage location", "stor. loc.", "stor loc",
                "storage loc", "sloc")).strip()

            if not article or article == "None":
                continue

            diag_count += 1
            if len(diag_sites) < 5:
                diag_sites.add(repr(site))
            if len(diag_slocs) < 5:
                diag_slocs.add(repr(storage_loc))

            if site_filter and site != site_filter:
                continue
            if storage_loc_filter and storage_loc != storage_loc_filter:
                continue

            clean = {}
            for k, v in row_dict.items():
                try:
                    clean[k] = float(v) if v is not None else ""
                except (TypeError, ValueError):
                    clean[k] = str(v).strip() if v is not None else ""

            if article in data:
                merge_row(data[article], clean)
            else:
                data[article] = clean

        wb.close()

        if (site_filter or storage_loc_filter) and not is_simple:
            if diag_count == 0:
                print(f"    SAP: no standard-format rows in {os.path.basename(path)}")
            elif not data:
                print(f"    SAP filter site='{site_filter}' "
                      f"sloc='{storage_loc_filter}' matched 0 of {diag_count} rows.")
                print(f"      Actual Site values:             {', '.join(sorted(diag_sites))}")
                print(f"      Actual Storage Location values: {', '.join(sorted(diag_slocs))}")
                print(f"      -> Update SAP_SITE / STORAGE_LOC constants to match.")
            else:
                print(f"    SAP filter site='{site_filter}' matched "
                      f"{len(data)} articles from {diag_count} rows.")

    except Exception as e:
        print(f"    SAP read error: {e}")

    return data, all_headers

def read_sap_files(paths, site_filter=None, storage_loc_filter=None):
    """Read and merge multiple SAP files. Returns (merged_data, headers)."""
    merged, all_headers = {}, []
    for p in paths:
        print(f"    SAP source: {os.path.basename(p)}")
        data, hdrs = read_sap_file(p, site_filter=site_filter,
                                   storage_loc_filter=storage_loc_filter)
        if not all_headers:
            all_headers = hdrs
        for article, row in data.items():
            if article in merged:
                merge_row(merged[article], row)
            else:
                merged[article] = dict(row)
    return merged, all_headers
