"""
Reader for user-modified ("Modify*") files.

Accepts both simple SKU/Stock format and full SAP-style columns.
Both formats use the SKU column as the lookup key.

Public API
----------
read_modify_file(path)               -> (data, headers)
read_modify_files(paths)             -> (merged_data, headers)
scan_modify_columns(paths)           -> list[str]
scan_modify_columns_per_file(paths)  -> {filepath: [col, ...]}
"""
import os

from ..utils import is_valid_sku, merge_row

def scan_modify_columns(modify_paths):
    """Return a deduplicated ordered header list from Modify files."""
    cols, seen = [], set()
    for path in modify_paths:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                raw_hdrs = [str(h).strip() if h else "" for h in row]
                hdrs = [h for h in raw_hdrs if h]
                if not hdrs:
                    continue
                for h in hdrs:
                    if h and h not in seen:
                        seen.add(h)
                        cols.append(h)
                break
            wb.close()
        except Exception as e:
            print(f"    Modify scan error {path}: {e}")
    return cols

def read_modify_file(path):
    """
    Read a single Modify file preserving original column names.
    Duplicate SKU values are merged by summing numeric fields.

    Returns (data, headers) where data is {sku: {col: val}}.
    """
    data, all_headers = {}, []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        headers = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                all_headers = [h for h in headers if h]
                continue
            if not any(row):
                continue

            row_dict = dict(zip(headers, row))
            hdrs_up = {h.upper(): h for h in headers if h}
            key_col = hdrs_up.get("SKU") or hdrs_up.get("ARTICLE") or headers[0]
            key_val = str(row_dict.get(key_col, "")).strip()
            if not key_val or key_val == "None" or not is_valid_sku(key_val):
                continue

            clean = {k: (str(v).strip() if v is not None else "")
                     for k, v in row_dict.items() if k}
            if key_val in data:
                merge_row(data[key_val], clean)
            else:
                data[key_val] = clean
        wb.close()
    except Exception as e:
        print(f"    Modify read error {path}: {e}")
    return data, all_headers

def read_modify_files(paths):
    """Read and merge multiple Modify files. Returns (merged_data, headers)."""
    merged, all_headers = {}, []
    for p in paths:
        print(f"    Modify source: {os.path.basename(p)}")
        data, hdrs = read_modify_file(p)
        if not all_headers:
            all_headers = hdrs
        for sku, row in data.items():
            if sku in merged:
                merge_row(merged[sku], row)
            else:
                merged[sku] = dict(row)
    return merged, all_headers

def scan_modify_columns_per_file(modify_paths):
    """Return columns per file as {filepath: [col, ...]}."""
    result = {}
    for path in modify_paths:
        cols = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                raw_hdrs = [str(h).strip() if h else "" for h in row]
                cols = [h for h in raw_hdrs if h]
                break
            wb.close()
        except Exception as e:
            print(f"    Modify scan error {path}: {e}")
        result[path] = cols
    return result
