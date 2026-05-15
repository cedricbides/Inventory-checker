"""
Auto-detection of source files from a working directory.

Exports
-------
find_files(base_dir)       -> channel_files, ordazzle_files, sap_files,
                              modify_files_list, modify_source, modify_files
sniff_file_source(path)    -> 'channel' | 'ordazzle' | 'sap' | None
is_user_modified_xlsx(path) -> bool
"""
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET

from .utils import get_shared_strings, xml_cell_val

MAX_ROWS_TO_SCAN = 30

from .config import FILE_PATTERNS as _FP

_SHOPEE_PFX       = tuple(_FP["shopee_prefixes"])
_LAZADA_PFX       = tuple(_FP["lazada_prefixes"])
_ZALORA_PFX       = tuple(_FP["zalora_prefixes"])
_ZALORA_KW        = tuple(_FP["zalora_keywords"])
_ZALORA_CSV_PFX   = tuple(_FP.get("zalora_csv_prefixes", ["zalorastock"]))
_ORDAZZLE_PFX     = tuple(_FP["ordazzle_prefixes"])
_SAP_PFX          = tuple(_FP["sap_prefixes"])
_MODIFY_PFX       = tuple(_FP["modify_prefixes"])

SAP_SYSTEM_INDICATORS = [
    "sap", "system", "service", "robot", "bot", "admin",
    "export", "report", "batch", "etl", "integration",
]

def is_user_modified_xlsx(path):
    """
    Return True when docProps/core.xml shows a real person (not a SAP process)
    as the last modifier.
    """
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext not in (".xlsx", ".xls"):
            return False
        with zipfile.ZipFile(path) as z:
            if "docProps/core.xml" not in z.namelist():
                return False
            with z.open("docProps/core.xml") as f:
                content = f.read().decode("utf-8", errors="replace")
        last_mod = re.search(r"<cp:lastModifiedBy>(.*?)</cp:lastModifiedBy>", content)
        creator = re.search(r"<dc:creator>(.*?)</dc:creator>", content)
        names = [
            (last_mod.group(1).strip() if last_mod else ""),
            (creator.group(1).strip() if creator else ""),
        ]
        for name in names:
            if name and not any(ind in name.lower() for ind in SAP_SYSTEM_INDICATORS):
                return True
    except Exception:
        pass
    return False

def _classify_row(cells):
    """Return source label for a set of header cell values, or None."""
    if "Product ID" in cells and (
        "SKU" in cells or "SellerSKU" in cells or "sku.skuId" in cells
    ):
        return "channel"
    if "PRODUCT CODE" in cells:
        return "ordazzle"
    if "Article" in cells and any(
        kw in cells for kw in ("Site", "Unrestricted Stock", "Unrestricted",
                               "Storage Location", "Stor. Loc.")
    ):
        return "sap"
    cells_up = {str(c).strip().upper() for c in cells}
    if "SKU" in cells_up and "STOCK" in cells_up and "PRODUCT ID" not in cells_up:
        return "sap"
    return None

def _sniff_xlsx_bytes(content):
    """Scan rows of an xlsx blob and return detected source, or None."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as inner:
            strings = get_shared_strings(inner)
            sheet_names = [n for n in inner.namelist()
                           if n.startswith("xl/worksheets/sheet")]
            for sheet_path in sorted(sheet_names):
                with inner.open(sheet_path) as ws:
                    root = ET.parse(ws).getroot()
                ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                row_count = 0
                for row in root.findall(".//x:row", ns):
                    cells = {xml_cell_val(c, strings, ns)
                             for c in row.findall("x:c", ns)}
                    cells.discard("")
                    if not cells:
                        continue
                    result = _classify_row(cells)
                    if result:
                        return result
                    row_count += 1
                    if row_count >= MAX_ROWS_TO_SCAN:
                        break
    except Exception as e:
        print(f"    sniff parse error: {e}")
    return None

def _sniff_openpyxl(path):
    """Fallback sniff via openpyxl (handles edge-case xlsx encodings)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = {str(v).strip() for v in row if v is not None and str(v).strip()}
                if not cells:
                    continue
                result = _classify_row(cells)
                if result:
                    wb.close()
                    return result
                if i >= MAX_ROWS_TO_SCAN:
                    break
        wb.close()
    except Exception as e:
        print(f"    sniff openpyxl error: {e}")
    return None

def sniff_file_source(file_path):
    """
    Determine whether a file is a 'channel', 'ordazzle', or 'sap' source.
    Returns None if undetermined.
    """
    ext = os.path.splitext(file_path)[1].lower()
    fname_low = os.path.basename(file_path).lower()
    src = None

    try:
        if ext == ".zip":
            with zipfile.ZipFile(file_path) as outer:
                xlsx_files = sorted(n for n in outer.namelist() if n.endswith(".xlsx"))
                if xlsx_files:
                    with outer.open(xlsx_files[0]) as f:
                        src = _sniff_xlsx_bytes(f.read())
        elif ext in (".xlsx", ".xls"):
            with open(file_path, "rb") as f:
                src = _sniff_xlsx_bytes(f.read())
            if src is None:
                src = _sniff_openpyxl(file_path)

        if src is None:
            if any(kw in fname_low for kw in ("mass", "shopee", "lazada", "price", "zalora")):
                src = "channel"
                print(f"    sniff fallback: filename keyword -> treating as Channel")
            elif any(kw in fname_low for kw in ("exp", "exl", "ordazzle", "buffer")):
                src = "ordazzle"
                print(f"    sniff fallback: filename keyword -> treating as Ordazzle")

        return src

    except Exception as e:
        print(f"    sniff error for {os.path.basename(file_path)}: {e}")
    return None

def find_files(base_dir=None):
    """
    Walk base_dir (defaults to cwd) and sort every file into one of four buckets
    by naming convention and/or metadata.

    Returns
    -------
    channel_files    : list[str]
    ordazzle_files   : list[str]
    sap_files        : list[str]
    modify_files_list: list[str]
    modify_source    : str | None
    modify_files     : list[(str, str)]
    """
    if base_dir is None:
        base_dir = os.getcwd()

    channel_raw, ordazzle_raw, sap_raw, modify_raw = [], [], [], []
    modify_source = None
    modify_files = []

    SKIP_DIRS = {"result", "__pycache__", ".git"}

    for root_dir, dirs, files in os.walk(base_dir):
        dirs[:] = [d for d in dirs if d.lower() not in SKIP_DIRS]

        for name in files:
            low = name.lower()
            full = os.path.join(root_dir, name)

            name_is_modify = low.startswith(_MODIFY_PFX) and low.endswith(
                (".zip", ".xlsx", ".xls")
            )
            metadata_is_user = (
                not name_is_modify
                and low.endswith((".xlsx", ".xls"))
                and not low.startswith(_SHOPEE_PFX + _LAZADA_PFX + _ZALORA_PFX + _ZALORA_CSV_PFX + _ORDAZZLE_PFX + _SAP_PFX + _MODIFY_PFX + ("inventory_result",))
                and "sellerstocktemplate" not in low
                and is_user_modified_xlsx(full)
            )

            if name_is_modify or metadata_is_user:
                reason = "Modify* filename" if name_is_modify else "user-modified metadata"
                detected = sniff_file_source(full)
                print(f"  User-modified file ({reason}): '{name}' -> sniffed as '{detected}'")
                modify_raw.append(full)
                modify_source = "modify"
                modify_files.append((full, detected or "modify"))

            elif low.startswith(_SHOPEE_PFX) and low.endswith((".zip", ".xlsx")):
                channel_raw.append(full)
            elif low.startswith(_LAZADA_PFX) and low.endswith((".xlsx", ".xls")):
                channel_raw.append(full)
            elif low.startswith(_ZALORA_PFX) and low.endswith((".xlsx", ".xls", ".zip")):
                channel_raw.append(full)
            elif any(kw in low for kw in _ZALORA_KW) and low.endswith((".xlsx", ".xls", ".zip")):
                channel_raw.append(full)
            elif low.startswith(_ZALORA_CSV_PFX) and low.endswith(".csv"):
                channel_raw.append(full)
            elif low.startswith(_ORDAZZLE_PFX) and low.endswith((".xls", ".xlsx")):
                ordazzle_raw.append(full)
            elif low.startswith(_SAP_PFX) and low.endswith((".xlsx", ".xls")):
                sap_raw.append(full)

    def _dedupe(lst):
        seen, out = set(), []
        for f in sorted(lst, key=str.lower):
            if f.lower() not in seen:
                seen.add(f.lower())
                out.append(f)
        return out

    return (
        _dedupe(channel_raw),
        _dedupe(ordazzle_raw),
        _dedupe(sap_raw),
        _dedupe(modify_raw),
        modify_source,
        modify_files,
    )