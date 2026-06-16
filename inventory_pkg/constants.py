"""
constants.py

Derived constants built from config.py.  Modules import what they need here
rather than re-deriving it themselves.
"""

from openpyxl.styles import PatternFill
from .config import BRAND_GROUPS, BRAND_ALIASES

# ── Flat brand list and keyword map ────────────────────────────────────────────

ALL_BRANDS = [brand for group in BRAND_GROUPS for brand in group["brands"]]

# Maps lowercased brand/alias strings to their canonical brand name.
# Longer keywords are checked first (see detect_brand_from_name in utils.py).
BRAND_KEYWORD_MAP = {b.lower(): b for b in ALL_BRANDS}
BRAND_KEYWORD_MAP.update({k.lower(): v for k, v in BRAND_ALIASES.items()})

# ── Per-group brand lists (used in utils.brand_group) ─────────────────────────

def _brands_for(group_name):
    for group in BRAND_GROUPS:
        if group["group"] == group_name:
            return group["brands"]
    return []

SSIEBG_BRANDS = _brands_for("SSIEBG")
PAYLESS_BRANDS = _brands_for("PAYLESS")
SLCI_BRANDS    = _brands_for("SLCI")


# ── Excel fill colours ─────────────────────────────────────────────────────────

GREEN      = PatternFill("solid", fgColor="C6EFCE")
RED        = PatternFill("solid", fgColor="FFC7CE")
YELLOW     = PatternFill("solid", fgColor="FFEB9C")
HEADER     = PatternFill("solid", fgColor="1A2A4A")
GRAY       = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
ORANGE     = PatternFill("solid", fgColor="FFD966")
CMP_HDR    = PatternFill("solid", fgColor="2E4A7A")


# ── Result Preview sheet static data ──────────────────────────────────────────

RESULT_PREVIEW_HEADERS = [
    "Base_SKU (Ordazzle)", "BASE_QTY (Ordazzle)", "CH_PRODUCT ID",
    "CH_Variation ID", "ORD_Inventory published", "CH_Stock",
    "SAP_UNRESTRICTED STOCK", "Ordazzle X Channel",
]

RESULT_PREVIEW_SAMPLE = [
    (1000048513,    1, "#N/A", "#N/A", "#N/A", "#N/A", "#N/A", "n/a"),
    (1000048706,    1, "58157771154", "435690761564", 0,  0,  1,  "TRUE"),
    (2000047394003, 1, "58157771155", "435690761565", 21, 21, 18, "TRUE"),
    (1000048845005, 1, "#N/A", "#N/A", "#N/A", "#N/A", "#N/A", "n/a"),
]

RESULT_PREVIEW_WIDTHS = [22, 20, 18, 18, 24, 12, 24, 20]