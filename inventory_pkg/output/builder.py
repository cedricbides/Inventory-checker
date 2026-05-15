"""
Excel output builder.

Public API
----------
build_output(brand, channel_data, ordazzle_data, sap_data, ...)
    -> (wb: Workbook, total_skus: int, matched_ord: int, matched_sap: int)
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..constants import (
    GREEN, RED, YELLOW, HEADER, GRAY, WHITE_FILL, ORANGE, CMP_HDR,
    RESULT_PREVIEW_HEADERS, RESULT_PREVIEW_SAMPLE, RESULT_PREVIEW_WIDTHS,
)
from ..utils import fmt_number, is_valid_sku

def _write_preview_row(ws, row_num, row_data, border, sample=False, ord_x_ch="n/a"):
    """Write one row to the Result Preview sheet with appropriate styling."""
    thin = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    SAMPLE_FILL  = PatternFill("solid", fgColor="F2F2F2")
    REAL_FILL    = (PatternFill("solid", fgColor="FFFFFF") if row_num % 2 == 0
                    else PatternFill("solid", fgColor="EEF4FB"))
    NA_FILL      = PatternFill("solid", fgColor="FFC7CE")
    TRUE_FILL    = PatternFill("solid", fgColor="C6EFCE")
    FALSE_FILL   = PatternFill("solid", fgColor="FFC7CE")
    NA_CMP_FILL  = PatternFill("solid", fgColor="FFEB9C")

    base_fill = SAMPLE_FILL if sample else REAL_FILL

    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font      = Font(size=9, color="999999" if sample else "222222", italic=sample)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

        if col_idx == len(row_data):
            oc = str(val).strip().upper()
            if oc == "TRUE":
                cell.fill = TRUE_FILL
                cell.font = Font(size=9, bold=True, color="276221")
            elif oc == "FALSE":
                cell.fill = FALSE_FILL
                cell.font = Font(size=9, bold=True, color="9C0006")
            else:
                cell.fill = NA_CMP_FILL
                cell.font = Font(size=9, color="7B5E00", italic=sample)
        elif str(val).strip() == "#N/A":
            cell.fill = NA_FILL
            cell.font = Font(size=9, color="9C0006", italic=True)
        else:
            cell.fill = base_fill

    ws.row_dimensions[row_num].height = 16

def _build_result_preview_sheet(wb, channel_data, ordazzle_data, sap_data, channel_label):
    """
    Add a 'Result Preview' sheet to *wb* that mirrors the sample xlsx VLOOKUP layout.

    - Real Ordazzle SKUs appear first (capped at 500 rows).
    - Static sample rows (greyed) are appended at the end.
    """
    ws = wb.create_sheet(title="Result Preview")

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(RESULT_PREVIEW_HEADERS))
    tc = ws.cell(row=1, column=1,
                 value=f"RESULT PREVIEW  |  Base: Ordazzle SKU  |  {channel_label}")
    tc.fill      = PatternFill("solid", fgColor="1A2A4A")
    tc.font      = Font(bold=True, color="FFFFFF", size=11)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Note row
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(RESULT_PREVIEW_HEADERS))
    note = ws.cell(row=2, column=1,
                   value=("ℹ  Preview: Base SKU iterates from Ordazzle. "
                          "Real data rows appear first; sample rows (grey) show expected format. "
                          "SKUs must be 10–13 digits starting with 1 or 2."))
    note.fill      = PatternFill("solid", fgColor="FFF8E7")
    note.font      = Font(italic=True, size=8, color="7B4F00")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Dynamic header (inject channel label)
    dynamic_headers    = list(RESULT_PREVIEW_HEADERS)
    dynamic_headers[2] = f"CH_PRODUCT ID ({channel_label})"
    dynamic_headers[5] = f"CH_Stock ({channel_label})"
    for col_idx, h in enumerate(dynamic_headers, 1):
        cell           = ws.cell(row=3, column=col_idx, value=h)
        cell.fill      = HEADER
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = ws.cell(row=4, column=1)

    # Real data rows
    real_rows_written = 0
    for sku in list(ordazzle_data.keys())[:500]:
        if not is_valid_sku(sku):
            continue
        ord_row = ordazzle_data.get(sku, {})
        ch_row  = channel_data.get(sku, {})
        sap_row = sap_data.get(sku, {})

        base_qty = ord_row.get("BASE QTY", ord_row.get("BASE_QTY", 1))
        try:
            base_qty = int(float(base_qty)) if base_qty not in ("", None) else 1
        except (ValueError, TypeError):
            base_qty = 1

        ch_product_id = ch_row.get("Product ID",   "#N/A") if ch_row else "#N/A"
        ch_variation  = ch_row.get("Variation ID",
                        ch_row.get("sku.skuId", "#N/A")) if ch_row else "#N/A"
        ord_inv_pub   = ord_row.get("INV TO PUBLISHED STOCK",
                        ord_row.get("PUBLISHED STOCK",
                        ord_row.get("Inventory published", "#N/A")))
        ch_stock      = ch_row.get("Stock", ch_row.get("Available Stock", "#N/A")) if ch_row else "#N/A"
        sap_unr       = sap_row.get("Unrestricted Stock",
                        sap_row.get("Unrestricted", "#N/A")) if sap_row else "#N/A"

        try:
            ord_x_ch = "TRUE" if float(ord_inv_pub) == float(ch_stock) else "FALSE"
        except (ValueError, TypeError):
            ord_x_ch = "n/a"

        row_data = [sku, base_qty, ch_product_id, ch_variation,
                    fmt_number(ord_inv_pub), fmt_number(ch_stock),
                    fmt_number(sap_unr), ord_x_ch]
        _write_preview_row(ws, 4 + real_rows_written, row_data, border,
                           sample=False, ord_x_ch=ord_x_ch)
        real_rows_written += 1

    # Sample rows
    for i, sample_row in enumerate(RESULT_PREVIEW_SAMPLE):
        _write_preview_row(ws, 4 + real_rows_written + i,
                           list(sample_row), border,
                           sample=True, ord_x_ch=str(sample_row[-1]))

    for col_idx, width in enumerate(RESULT_PREVIEW_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _build_quick_export_sheet(wb, channel_data, ordazzle_data, channel, channel_label):
    """
    Add a 'Quick Export' sheet to *wb* with a fixed focused layout:

        SKU | Stock | INV TO PUBLISHED STOCK | Product ID | Variation ID
            | [CH·??] Stock × [ORD] INV TO PUBLISHED STOCK

    Iterates over Ordazzle SKUs so every warehouse row is represented.
    """
    CH_ABBR = {
        "Channel_Shopee": "CH·SP",
        "Channel_Lazada": "CH·LZ",
        "Channel_Zalora": "CH·ZL",
    }.get(channel_label, "CH")

    ORD_INV_KEYS = ("INV TO PUBLISHED STOCK", "PUBLISHED STOCK", "Inventory published")
    CH_STOCK_KEYS = ("Stock", "Available Stock", "quantity", "Quantity")
    CH_VAR_KEYS   = ("Variation ID", "sku.skuId", "SKU Reference No.")

    cmp_label = f"[{CH_ABBR}] Stock × [ORD] INV TO PUBLISHED STOCK"
    HEADERS    = ["SKU", "Stock", "INV TO PUBLISHED STOCK",
                  "Product ID", "Variation ID", cmp_label]
    COL_WIDTHS = [18, 10, 26, 18, 18, 44]

    ws = wb.create_sheet(title="Quick Export")

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(HEADERS))
    tc = ws.cell(row=1, column=1,
                 value=f"QUICK EXPORT  |  Ordazzle × {channel}  |  {channel_label}")
    tc.fill      = PatternFill("solid", fgColor="1A2A4A")
    tc.font      = Font(bold=True, color="FFFFFF", size=11)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    CMP_HDR_FILL = PatternFill("solid", fgColor="2E4A7A")
    for col_idx, h in enumerate(HEADERS, 1):
        cell           = ws.cell(row=2, column=col_idx, value=h)
        cell.fill      = CMP_HDR_FILL if col_idx == len(HEADERS) else HEADER
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = ws.cell(row=3, column=1)

    NA_FILL = PatternFill("solid", fgColor="FFC7CE")

    row_num = 3
    for sku, ord_row in ordazzle_data.items():
        if not is_valid_sku(sku):
            continue

        ch_row = channel_data.get(sku, {})

        # Resolve ordazzle INV TO PUBLISHED STOCK
        ord_inv = "#N/A"
        for k in ORD_INV_KEYS:
            v = ord_row.get(k)
            if v not in ("", None):
                ord_inv = v
                break

        # Resolve channel columns (empty dict if SKU not in channel)
        if ch_row:
            ch_stock = "#N/A"
            for k in CH_STOCK_KEYS:
                v = ch_row.get(k)
                if v not in ("", None):
                    ch_stock = v
                    break
            product_id   = ch_row.get("Product ID", "#N/A") or "#N/A"
            variation_id = "#N/A"
            for k in CH_VAR_KEYS:
                v = ch_row.get(k)
                if v not in ("", None):
                    variation_id = v
                    break
        else:
            ch_stock = product_id = variation_id = "#N/A"

        # Comparison
        try:
            cmp_val = "TRUE" if float(ord_inv) == float(ch_stock) else "FALSE"
        except (ValueError, TypeError):
            cmp_val = "N/A"

        fill_base = GRAY if row_num % 2 == 0 else WHITE_FILL

        row_values = [
            sku,
            fmt_number(ch_stock),
            fmt_number(ord_inv),
            product_id,
            variation_id,
            cmp_val,
        ]
        for col_idx, val in enumerate(row_values, 1):
            cell           = ws.cell(row=row_num, column=col_idx,
                                     value=val if val != "" else None)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

            if col_idx == len(HEADERS):           # comparison column
                if val == "TRUE":
                    cell.fill = GREEN
                    cell.font = Font(size=9, bold=True, color="276221")
                elif val == "FALSE":
                    cell.fill = RED
                    cell.font = Font(size=9, bold=True, color="9C0006")
                else:
                    cell.fill = YELLOW
                    cell.font = Font(size=9, color="7B5E00")
            elif str(val) == "#N/A":              # missing lookup
                cell.fill = NA_FILL
                cell.font = Font(size=9, color="9C0006", italic=True)
            else:
                cell.fill = fill_base
                cell.font = Font(size=9)

        ws.row_dimensions[row_num].height = 15
        row_num += 1

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def build_output(
    brand, channel_data, ordazzle_data, sap_data,
    channel="Shopee",   
    channel_label=None,
    ordazzle_warning=None,
    sap_warning=None,
    base_sku="channel",
    selected_output_cols=None,
    comparisons=None,
    modify_data=None,
):
    """
    Build the final inventory comparison workbook.

    Parameters
    ----------
    brand                : str
    channel_data         : dict  {sku: {col: val}}
    ordazzle_data        : dict  {sku: {col: val}}
    sap_data             : dict  {sku: {col: val}}
    channel              : 'Shopee' | 'Lazada' | 'Zalora'
    channel_label        : str | None  – e.g. 'Channel_Shopee'
    ordazzle_warning     : str | None
    sap_warning          : str | None
    base_sku             : 'channel' | 'ordazzle' | 'sap' | 'modify'
    selected_output_cols : list[(source, col_name)]
    comparisons          : list[(lsrc, lcol, rsrc, rcol, label)]
    modify_data          : dict | None

    Returns
    -------
    wb            : openpyxl.Workbook
    total_skus    : int
    matched_ord   : int
    matched_sap   : int
    """
    if selected_output_cols is None:
        selected_output_cols = []
    if comparisons is None:
        comparisons = []
    if modify_data is None:
        modify_data = {}
    if channel_label is None:
        channel_label = {
            "Lazada": "Channel_Lazada",
            "Zalora": "Channel_Zalora",
            "Shopee": "Channel_Shopee",
        }.get(channel, "Channel_Shopee")

    # Drop columns that belong to a different channel type
    _active_ch_src = {
        "Lazada": "channel_lazada",
        "Shopee": "channel_shopee",
        "Zalora": "channel_zalora",
    }.get(channel, None)

    def _ch_ok(src):
        if not src.startswith("channel_"):
            return True
        if _active_ch_src is None:
            return True
        return src == _active_ch_src

    selected_output_cols = [(s, c) for s, c in selected_output_cols if _ch_ok(s)]
    comparisons = [
        (ls, lc, rs, rc, lb) for ls, lc, rs, rc, lb in comparisons
        if _ch_ok(ls) and _ch_ok(rs)
    ]

    # Enforce canonical Quick Export column order when all selected columns
    # match the QE set.  This fixes multi-channel sessions where the shared
    # Ordazzle column ends up before the channel columns for the 2nd marketplace.
    _QE_ORDER_IDX = {
        "SKU": 0, "SellerSKU": 0, "Parent SKU": 0, "seller_sku": 0,
        "Stock": 1, "Available Stock": 1, "quantity": 1, "Quantity": 1,
        "INV TO PUBLISHED STOCK": 2, "PUBLISHED STOCK": 2, "Inventory published": 2,
        "Product ID": 3,
        "Variation ID": 4, "sku.skuId": 4, "SKU Reference No.": 4,
    }
    _qe_set = set(_QE_ORDER_IDX)
    if selected_output_cols and all(col in _qe_set for _, col in selected_output_cols):
        selected_output_cols = sorted(
            selected_output_cols,
            key=lambda sc: _QE_ORDER_IDX.get(sc[1], 99),
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTORY RESULT"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_data    = len(selected_output_cols)
    n_cmp     = len(comparisons)
    total_cols = n_data + n_cmp or 1

    start_row = 1
    if ordazzle_warning or sap_warning:
        msg = " | ".join(filter(None, [ordazzle_warning, sap_warning]))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        cell           = ws.cell(row=1, column=1, value=msg)
        cell.fill      = ORANGE
        cell.font      = Font(bold=True, size=10, color="7B3F00")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36
        start_row = 2

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=total_cols)
    bc = ws.cell(row=start_row, column=1,
                 value=f"Brand: {brand}  |  {channel} Channel  [{channel_label}]")
    bc.fill      = PatternFill("solid", fgColor="2E4A7A")
    bc.font      = Font(bold=True, color="FFFFFF", size=11)
    bc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22
    start_row += 1

    SRC_CLR = {
        "channel": "2C6E8A", "channel_shopee": "2C6E8A",
        "channel_lazada": "1A5F7A", "channel_zalora": "0D4D6A",
        "ordazzle": "5A4A8A", "sap": "2A6A4A", "modify": "8B5E00",
    }
    SRC_NAMES = {
        "channel": channel, "channel_shopee": "Shopee",
        "channel_lazada": "Lazada", "channel_zalora": "Zalora",
        "ordazzle": "Ordazzle", "sap": "SAP", "modify": "Modify",
    }
    for i, (src, _) in enumerate(selected_output_cols, 1):
        cell           = ws.cell(row=start_row, column=i, value=SRC_NAMES.get(src, src))
        cell.fill      = PatternFill("solid", fgColor=SRC_CLR.get(src, "444444"))
        cell.font      = Font(bold=True, color="FFFFFF", size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    for i in range(n_cmp):
        cell           = ws.cell(row=start_row, column=n_data + i + 1, value="Comparison")
        cell.fill      = CMP_HDR
        cell.font      = Font(bold=True, color="FFFFFF", size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[start_row].height = 16
    start_row += 1

    header_row = start_row
    headers    = [col_name for _, col_name in selected_output_cols] + \
                 [lbl for _, _, _, _, lbl in comparisons]
    for i, h in enumerate(headers, 1):
        cell           = ws.cell(row=header_row, column=i, value=h)
        cell.fill      = HEADER
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    if base_sku == "ordazzle":
        iter_keys = list(ordazzle_data.keys())
    elif base_sku == "sap":
        iter_keys = list(sap_data.keys())
    elif base_sku == "modify":
        iter_keys = list(modify_data.keys())
    else:
        iter_keys = list(channel_data.keys())

    def get_val(src, col, ch_row, ord_row, sap_row, mod_row):
        if src.startswith("channel"):
            row = ch_row
        else:
            row = {"ordazzle": ord_row, "sap": sap_row, "modify": mod_row}.get(src, {})
        return row.get(col, "") if row else ""

    def cmp_result(a, b):
        if a == "" or b == "" or a is None or b is None:
            return "N/A"
        try:
            return "TRUE" if float(a) == float(b) else "FALSE"
        except (TypeError, ValueError):
            return "TRUE" if str(a).strip() == str(b).strip() else "FALSE"

    row_num          = header_row + 1
    matched_ordazzle = matched_sap = 0

    for sku in iter_keys:
        ch_row  = channel_data.get(sku, {})
        ord_row = ordazzle_data.get(sku, {})
        sap_row = sap_data.get(sku, {})
        mod_row = modify_data.get(sku, {})

        # Lazada: also try Product ID as lookup key
        if channel == "Lazada" and ch_row:
            pid = ch_row.get("Product ID", "").strip()
            if not ord_row and pid:
                ord_row = ordazzle_data.get(pid, {})
            if not sap_row and pid:
                sap_row = sap_data.get(pid, {})
            if not mod_row and pid:
                mod_row = modify_data.get(pid, {})

        if ord_row:
            matched_ordazzle += 1
        if sap_row:
            matched_sap += 1

        fill_base = GRAY if row_num % 2 == 0 else WHITE_FILL

        for col_idx, (src, col_name) in enumerate(selected_output_cols, 1):
            val  = get_val(src, col_name, ch_row, ord_row, sap_row, mod_row)
            val  = fmt_number(val)
            cell = ws.cell(row=row_num, column=col_idx, value=val if val != "" else None)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            cell.fill      = fill_base

        for i, (lsrc, lcol, rsrc, rcol, lbl) in enumerate(comparisons):
            lv     = get_val(lsrc, lcol, ch_row, ord_row, sap_row, mod_row)
            rv     = get_val(rsrc, rcol, ch_row, ord_row, sap_row, mod_row)
            result = cmp_result(lv, rv)
            cell   = ws.cell(row=row_num, column=n_data + i + 1, value=result)
            cell.font      = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            cell.fill      = GREEN if result == "TRUE" else (RED if result == "FALSE" else YELLOW)

        row_num += 1

    for col_idx in range(1, total_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(row_num, 60))),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)

    total = row_num - header_row - 1
    return wb, total, matched_ordazzle, matched_sap